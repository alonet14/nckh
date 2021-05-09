import pandas
from scipy.signal import butter
from scipy.signal import lfilter
from matplotlib import pyplot as plt
from datetime import datetime

import openpyxl
import scipy.signal as ss
import calendar
import pytz

fs = 100
fL = 0.83
fH = 2.33
fr = 0.5


def butter_bandpass_filter(data, lowcut, highcut, fs, order=3):
    nyq = 0.5 * fs
    low = lowcut / nyq
    high = highcut / nyq

    b, a = butter(order, [low, high], btype='band')
    y = lfilter(b, a, data)
    return y
################################### get data #################################
wb = openpyxl.load_workbook("rate-100-NofS-100-60s.xlsx")
sheet = wb.sheetnames

print(sheet)

sh1 = wb[sheet[0]]
row = sh1.max_row
print(row)

######### get voltage #######
voltage = [None] * (row - 1)
for index in range(2, row + 1):
    voltage[index - 2] = sh1.cell(row=index, column=2).value
print(voltage)
######### get time ###########
times = [0] * (row - 1)
timestamp = [0] * (row - 1)
for index in range(2, row + 1):
    times[index - 2] = sh1.cell(row=index, column=1).value

    chicago_tz = pytz.timezone('America/Chicago')
    chicago_dt = chicago_tz.localize(times[index - 2])

    timestamp[index - 2] = calendar.timegm(chicago_dt.utctimetuple())

print(timestamp)

################################### draw data #################################
plt.figure(1)
plt.title('Original Signal')
plt.xlabel('Times')
plt.ylabel('Voltage')
plt.plot(times, voltage)

plt.figure(2)
HR = butter_bandpass_filter(voltage, fL, fH, fs)
plt.title('HR signal')
plt.xlabel('Times')
plt.ylabel('Voltage')
plt.plot(times, HR)

threshold_HR = (max(HR) - min(HR)) * 0.01
peaks = ss.find_peaks(HR, threshold=threshold_HR, height=1.35, distance=1.5)

# plt.plot(timestamp[peaks], HR[peaks], 'x')


plt.show()
